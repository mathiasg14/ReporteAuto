import openpyxl

import pandas as pd
import dataframe_image as dfi
import os
import matplotlib.pyplot as plt
import numpy as np

from datetime import datetime
from pathlib import Path


def gen_reporte(file_name, mes_inicio, mes_fin, num_pol=None):
    # Extraer Pestañas de Reporte de Siniestralidad para Generación de Resumenes
    if type(num_pol) is not list and num_pol is not None:
        num_pol = [num_pol]
    # Abrir siniestralidad como openyxl object
    ab_path = Path(os.getcwd())
    rel_path = Path(f'../docs/company/{file_name}_Siniestralidad.xlsx')
    sin_doc = openpyxl.load_workbook(filename=rel_path, data_only=True)

    # Abrir detalles como openpyxl object
    rel_path = Path(f'../docs/company/{file_name}_Detalle.xlsx')
    det_doc = openpyxl.load_workbook(filename=rel_path, data_only=True)

    # Retornar pestanas correspondientes del Excel Siniestralidad

    if num_pol is not None:
        pest = [sin_doc['Grupos'].cell(row, 1).value for row in
                [i for i in range(5, sin_doc['Grupos'].max_row) if
                 sin_doc['Grupos'].cell(i, 2).value in num_pol]]
    else:
        pest = [sin_doc['Grupos'].cell(row, 1).value for row in range(5, sin_doc['Grupos'].max_row)]

    # Generar carpeta de reporte
    report_date = datetime.now()
    report_date = report_date.strftime('%y-%m-%d-%H-%M')
    route_path = Path(f"../docs/reportes/reporte_{report_date}/")
    os.mkdir(route_path)

    # Generar tabla de salud y vida
    for p in pest:
        num_pol = sin_doc['Grupos'].cell(4 + p, 2).value
        contratante = sin_doc['Grupos'].cell(4 + p, 3).value

        # convertir en DF informacion de tabla principal (mes x mes)
        ltm_data = []

        for index, val in enumerate(sin_doc[str(p)]['A7:G18']):
            s_list = []
            for sub_ind, sub_val in enumerate(val):
                s_list.append(sub_val.value)
            ltm_data.append(s_list)

        df_ltm = pd.DataFrame(ltm_data)
        headers = ['Periodo', 'Prima Vida', 'Prima Salud', 'Reclamo Vida', 'Reclamo Salud',
                   'Siniestralidad Vida', 'Siniestralidad Salud']
        df_ltm.columns = headers

        df_range = df_ltm.loc[(df_ltm['Periodo'] >= f'{mes_inicio[1]}-{mes_inicio[0]}-01')
                              & (df_ltm['Periodo'] <= f'{mes_fin[1]}-{mes_fin[0]}-01')]

        # Exportación de Tablas de Salud a imagenes
        df_salud = df_range.iloc[:, [0, 2, 4, 6]]
        df_salud.iloc[:, 0] = df_salud.iloc[:, 0].dt.strftime('%m-%y')
        df_salud.iloc[:, 1] = df_salud.iloc[:, 1]
        df_salud.iloc[:, 2] = df_salud.iloc[:, 2]
        df_salud.iloc[:, 3] = df_salud.iloc[:, 3]


        mes_tag = f'{mes_inicio[1]}_{mes_inicio[0]}-{mes_fin[1]}_{mes_fin[0]}'

        save_path = Path(f"{route_path}/{mes_tag}_{num_pol}_salud.png")
        dfi.export(df_salud, save_path)

        fig, ax = plt.subplots(figsize=(20,10))


        df_salud.plot(x='Periodo',y=['Prima Salud','Reclamo Salud'],kind='bar', ax=ax)
        df_salud.plot(x='Periodo', y= 'Siniestralidad Salud', color = 'gray', kind='line', ax=ax, secondary_y=True)

        plt.show()

        # Exportación de Tablas de Vida a imagenes
        df_vida = df_range.iloc[:, [0, 1, 3, 5]]
        df_vida.iloc[:, 0] = df_vida.iloc[:, 0].dt.strftime('%m-%y')
        df_vida.iloc[:, 1] = df_vida.iloc[:, 1].apply(lambda x: f"${x:,.2f}")
        df_vida.iloc[:, 2] = df_vida.iloc[:, 2].apply(lambda x: f"${x:,.2f}")
        df_vida.iloc[:, 3] = df_vida.iloc[:, 3].apply(lambda x: f"{x * 100:,.2f}%")
        df_vida.set_index('Periodo', inplace=True)

        mes_tag = f'{mes_inicio[1]}_{mes_inicio[0]}-{mes_fin[1]}_{mes_fin[0]}'

        save_path = Path(f"{route_path}/{mes_tag}_{num_pol}_vida.png")
        dfi.export(df_vida, save_path)

        # Generación Tabla Tipo de Reclamante
        sum_aseg = 0
        sum_dep = 0
        mesini_dt = datetime.strptime(f'{str(mes_inicio[1])}-0{str(mes_inicio[0])}-01', '%Y-%m-%d')
        mesfin_dt = datetime.strptime(f'{str(mes_fin[1])}-0{str(mes_fin[0])}-01', '%Y-%m-%d')

        for row in range(5, det_doc['Aseg. Dep.'].max_row):
            if det_doc['Aseg. Dep.'].cell(row, 1).value == num_pol:
                if mesini_dt <= det_doc['Aseg. Dep.'].cell(row, 3).value <= mesfin_dt:
                    if det_doc['Aseg. Dep.'].cell(row, 6).value == 'Asegurado Principal':
                        sum_aseg += det_doc['Aseg. Dep.'].cell(row, 7).value
                    else:
                        sum_dep += det_doc['Aseg. Dep.'].cell(row, 7).value

        sum_tot = sum_aseg + sum_dep
        per_aseg = sum_aseg / sum_tot
        per_dep = sum_dep / sum_tot

        df_rec = pd.DataFrame({'Tipo de Reclamante': ['Asegurado', 'Dependiente', 'Total'],
                               'Reclamos Salud': [sum_aseg, sum_dep, sum_tot],
                               '% Util': [per_aseg * 100, per_dep * 100, 100]})

        df_rec_styled = pd.DataFrame({'Tipo de Reclamante': ['Asegurado', 'Dependiente', 'Total'],
                                      'Reclamos Salud': [f'${sum_aseg:,.2f}', f'${sum_dep:,.2f}', f'${sum_tot:,.2f}'],
                                      '% Util': [f'{per_aseg * 100:,.2f}%', f'{per_dep * 100:,.2f}%', f'100.00%']})

        save_path = Path(f"{route_path}/{mes_tag}_{num_pol}_Reclamante.png")
        dfi.export(df_rec_styled, save_path)

        print(f'Num Pol {num_pol} Contratante: {contratante}'
              f'\nAsegurado: ${sum_aseg:,.2f}'
              f'\nDependientes: ${sum_dep:,.2f}'
              f'\nTotal: ${sum_tot:,.2f}')

        # Generación Tabla Rubros Utilizados
        rubros = []
        rubros_dict_list = []
        for row in range(5, det_doc['Rubros'].max_row):
            if det_doc['Rubros'].cell(row, 1).value == num_pol:
                if mesini_dt <= det_doc['Rubros'].cell(row, 3).value <= mesfin_dt:
                    if (rub := det_doc['Rubros'].cell(row, 4).value) not in rubros:
                        rubros_dict_list.append({'RUBRO': rub,
                                                 'ASEGURADO': det_doc['Rubros'].cell(row, 5).value,
                                                 'DEPENDIENTE': det_doc['Rubros'].cell(row, 6).value})
                        rubros.append(rub)
                    else:
                        for _, x_val in enumerate(rubros_dict_list):
                            if x_val['RUBRO'] == rub:
                                x_val['ASEGURADO'] += det_doc['Rubros'].cell(row, 5).value
                                x_val['DEPENDIENTE'] += det_doc['Rubros'].cell(row, 6).value

        df_rub = pd.DataFrame(rubros_dict_list)
        df_rub['TOTAL'] = df_rub['ASEGURADO'] + df_rub['DEPENDIENTE']
        df_rub['UTIL %'] = df_rub['TOTAL'] / df_rub['TOTAL'].sum() * 100
        df_rub.loc['Totales'] = df_rub[['ASEGURADO', 'DEPENDIENTE', 'TOTAL', 'UTIL %']].sum()

        df_rub_styled = df_rub.style.format({
            "ASEGURADO": "${:,.2f}",
            "DEPENDIENTE": "${:,.2f}",
            "TOTAL": "${:,.2f}",
            "UTIL %": "{:,.2f}%"
        }).hide(axis="index")
        # print(f'\nNum Pol {num_pol} Contratante: {contratante}'
        #       f'\nRUBRO\tASEGURADO\tDEPENDIENTE\tTOTAL')
        # for val in rubros_dict_list:
        #     print(
        #         f'{val["RUBRO"]}\t{val["ASEGURADO"]:,.2f}\t{val["DEPENDIENTE"]:,.2f}\t{val["ASEGURADO"] + val["DEPENDIENTE"]:,.2f}')

        save_path = Path(f"{route_path}/{mes_tag}_{num_pol}_Rubros.png")
        dfi.export(df_rub_styled, save_path)

        # Generación de Diagnosticos Reportados
        diag = []
        diag_dict_list = []
        for row in range(5, det_doc['Dx.'].max_row):
            if det_doc['Dx.'].cell(row, 1).value == num_pol:
                if mesini_dt <= det_doc['Dx.'].cell(row, 3).value <= mesfin_dt:
                    if (dig := det_doc['Dx.'].cell(row, 4).value) not in diag:
                        diag_dict_list.append({'DIAGNOSTICO': dig,
                                               'MONTO': det_doc['Dx.'].cell(row, 5).value})
                        diag.append(dig)
                    else:
                        for _, x_val in enumerate(diag_dict_list):
                            if x_val['DIAGNOSTICO'] == dig:
                                x_val['MONTO'] += det_doc['Dx.'].cell(row, 5).value

        print(f'\nNum Pol {num_pol} Contratante: {contratante}'
              f'\nDIAGNOSTICO\tMONTO')
        for val in diag_dict_list:
            print(
                f'{val["DIAGNOSTICO"]}\t{val["MONTO"]:,.2f}')

        df_diag = pd.DataFrame(diag_dict_list)
        df_diag['%'] = df_diag['MONTO'] / df_diag['MONTO'].sum() * 100

        df_diag.loc['Totales'] = df_diag[['MONTO', '%']].sum()
        df_diag.loc['Totales', 'DIAGNOSTICO'] = 'TOTAL GENERAL'

        df_diag_styled = df_diag.style.format({
            "MONTO": "${:,.2f}",
            "%": "{:,.2f}%"
        }).hide(axis="index")
        # print(f'\nNum Pol {num_pol} Contratante: {contratante}'
        #       f'\nRUBRO\tASEGURADO\tDEPENDIENTE\tTOTAL')
        # for val in rubros_dict_list:
        #     print(
        #         f'{val["RUBRO"]}\t{val["ASEGURADO"]:,.2f}\t{val["DEPENDIENTE"]:,.2f}\t{val["ASEGURADO"] + val["DEPENDIENTE"]:,.2f}')

        save_path = Path(f"{route_path}/{mes_tag}_{num_pol}_Diagnosticos.png")
        dfi.export(df_diag_styled, save_path)


# def gen_barchart(df, filename):
#     # Create subplot and bar
#     fig, ax = plt.subplots()
#     x_axis = np.arange(len(df['PERIODO']))
#
#     ax.bar(x_axis - 0.2, df['Prima Salud'], width=0.4, label='Prima Salud')
#     ax.bar(x_axis + 0.2, df['Reclamo Salud'], width=0.4, label='Reclamo Salud')
#
#     ax.xticks(x_axis, df['PERIODO'])
#
#     ax.legend()
#
#     ax.show()
#
#     # headers = list(value for value in sin_doc["Grupos"].iter_rows(min_row=4, max_row=4, values_only=True))
#     # headers = list(headers[0])
#     #
#     # group_summary = []
#     #
#     # for index, val in enumerate(sin_doc["Grupos"][5:68]):
#     #     s_list = []
#     #     for sub_ind, sub_val in enumerate(val):
#     #         s_list.append(sub_val.value)
#     #     group_summary.append(s_list)
#     #
#     # df_resumen = pd.DataFrame(group_summary)
#     # df_resumen = df_resumen.loc[:, df_resumen.columns != 6]
#     # headers = [i.strip() for i in headers if i is not None]
#     # df_resumen.columns = headers
#     #
#     # pest = []
#     # if num_pol is not None:
#     #     for num in num_pol:
#     #         row = df_resumen.loc[df_resumen['Póliza'] == num].squeeze()
#     #         pest.append(row.get(key='Pestaña'))
#     # else:
#     #     pest = [sin_doc['Grupos'].cell(row, 1).value for row in range(5, sin_doc['Grupos'].max_row)]
#     #     print(pest)
#     # noinspection PyDictCreation
#     # totales_dict = {'Sin LTM': report_doc[str(pest)]['I19'].value,
#     #                 'LTM Reserva': report_doc[str(pest)]['K19'].value,
#     #                 'Sin Pol': report_doc[str(pest)]['I22'].value,
#     #                 'Pol Reserva': report_doc[str(pest)]['K22'].value,
#     #                 'Sin YTD': report_doc[str(pest)]['I25'].value,
#     #                 'YTD Reserva': report_doc[str(pest)]['K25'].value}

# def generar_tablas_reporte(file_name, mes_inicio, mes_fin, num_pol=None):
#     pestanas = pest_palig(file_name, num_pol)
#
#     return
#
#     file_path = Path(f"./docs/sample/{file_name}")
#     report_doc = openpyxl.load_workbook(filename=file_path, data_only=True)
#
#     report_date = datetime.now()
#     report_date = report_date.strftime('%y-%m-%d-%H-%M')
#     route_path = Path(f"./docs/reportes/reporte_{report_date}/")
#     os.mkdir(route_path)
#
#
#     for pest in pestanas:
#         num_pol = report_doc['Grupos'].cell(4+pest, 2).value
#         headers = list(value for value in report_doc[str(pest).strip()].iter_rows(min_row=6, max_row=6, values_only=True))
#         headers = list(headers[0])
#
#         # convertir en DF informacion de tabla principal (mes x mes)
#         ltm_data = []
#
#         for index, val in enumerate(report_doc[str(pest)]['A7:G18']):
#             s_list = []
#             for sub_ind, sub_val in enumerate(val):
#                 s_list.append(sub_val.value)
#             ltm_data.append(s_list)
#
#         df_ltm = pd.DataFrame(ltm_data)
#         headers = ['Periodo','Prima Vida', 'Prima Salud', 'Reclamo Vida', 'Reclamo Salud',
#                    'Siniestralidad Vida', 'Siniestralidad Salud']
#         df_ltm.columns = headers
#
#         df_range = df_ltm.loc[(df_ltm['Periodo'] >= f'{mes_inicio[1]}-{mes_inicio[0]}-01')
#                               & (df_ltm['Periodo'] <= f'{mes_fin[1]}-{mes_fin[0]}-01')]
#
#         # convertir en DF_2 informacion de Totales (Totales LTM / Sntr incurrida LTM / Reserva; Totales Año Pöliza,
#         # Totales YTD)
#
#         # noinspection PyDictCreation
#         totales_dict = {'Sin LTM': report_doc[str(pest)]['I19'].value, 'LTM Reserva': report_doc[str(pest)]['K19'].value,
#                         'Sin Pol': report_doc[str(pest)]['I22'].value, 'Pol Reserva': report_doc[str(pest)]['K22'].value,
#                         'Sin YTD': report_doc[str(pest)]['I25'].value, 'YTD Reserva': report_doc[str(pest)]['K25'].value}
#
#         # Generación de Tabla de Salud  en imagenes
#         df_salud = df_range.iloc[:, [0, 2, 4, 6]]
#         df_salud.iloc[:, 0] = df_salud.iloc[:, 0].dt.strftime('%m-%y')
#         df_salud.iloc[:, 1] = df_salud.iloc[:, 1].apply(lambda x: f"${x:,.2f}")
#         df_salud.iloc[:, 2] = df_salud.iloc[:, 2].apply(lambda x: f"${x:,.2f}")
#         df_salud.iloc[:, 3] = df_salud.iloc[:, 3].apply(lambda x: f"{x * 100:,.2f}%")
#         df_salud.set_index('Periodo', inplace=True)
#
#         mes_tag = f'{mes_inicio[1]}_{mes_inicio[0]}-{mes_fin[1]}_{mes_fin[0]}'
#
#         save_path = Path(f"{route_path}/{mes_tag}_{num_pol}_salud.png")
#         dfi.export(df_salud, save_path)
#
#         # Generación de Tabla de Salud  en imagenes
#         df_vida = df_range.iloc[:, [0, 1, 3, 5]]
#         df_vida.iloc[:, 0] = df_vida.iloc[:, 0].dt.strftime('%m-%y')
#         df_vida.iloc[:, 1] = df_vida.iloc[:, 1].apply(lambda x: f"${x:,.2f}")
#         df_vida.iloc[:, 2] = df_vida.iloc[:, 2].apply(lambda x: f"${x:,.2f}")
#         df_vida.iloc[:, 3] = df_vida.iloc[:, 3].apply(lambda x: f"{x * 100:,.2f}%")
#         df_vida.set_index('Periodo', inplace=True)
#
#         mes_tag = f'{mes_inicio[1]}_{mes_inicio[0]}-{mes_fin[1]}_{mes_fin[0]}'
#
#         save_path = Path(f"{route_path}/{mes_tag}_{num_pol}_vida.png")
#         dfi.export(df_vida, save_path)


# # identificar meses y años de cada fila
# meses = pd.Series(df_ltm['Periodo'].dt.month).tolist()
# df_ltm['MES'] = meses
#
# yrs = pd.Series(df_ltm['Periodo'].dt.year).tolist()
# df_ltm['YEAR'] = yrs
#
# mask = (int(mes_inicio) <= int(list(df_ltm['MES'])) <= int(mes_fin))
# df_range = df.loc[mask]

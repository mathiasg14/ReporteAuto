import openpyxl

import pandas as pd
import dataframe_image as dfi
import matplotlib.pyplot as plt
import numpy as np

from pathlib import Path


def read_report(file_name):
    file_path = Path(f"./docs/sample/{file_name}")
    report_doc = openpyxl.load_workbook(filename=file_path, data_only=True)

    headers = list(value for value in report_doc["Grupos"].iter_rows(min_row=4, max_row=4, values_only=True))
    headers = list(headers[0])

    group_summary = []

    for index, val in enumerate(report_doc["Grupos"][5:68]):
        s_list = []
        for sub_ind, sub_val in enumerate(val):
            s_list.append(sub_val.value)
        group_summary.append(s_list)

    df_grupos_resumen = pd.DataFrame(group_summary)
    df_grupos_resumen = df_grupos_resumen.loc[:, df_grupos_resumen.columns != 6]
    headers = [i.strip() for i in headers if i is not None]
    df_grupos_resumen.columns = headers

    return df_grupos_resumen


def resumen_grupo(file_name, num_pol, mes_inicio, mes_fin):
    # if type(num_pol) is not list:
    #     num_pol = [num_pol]

    df = read_report(file_name)
    polizas = list(df['Póliza'])

    # Find numero de pestaña from DF
    row = df.loc[df['Póliza'] == num_pol].squeeze()
    pest = row.get(key='Pestaña')

    # Abrir workbook y hoja correspondiente a número de pestaña
    file_path = Path(f"./docs/sample/{file_name}")
    report_doc = openpyxl.load_workbook(filename=file_path, data_only=True)

    headers = list(value for value in report_doc[str(pest).strip()].iter_rows(min_row=6, max_row=6, values_only=True))
    headers = list(headers[0])

    # convertir en DF informacion de tabla principal (mes x mes)
    ltm_data = []

    for index, val in enumerate(report_doc[str(pest)]['A7:G18']):
        s_list = []
        for sub_ind, sub_val in enumerate(val):
            s_list.append(sub_val.value)
        ltm_data.append(s_list)

    df_ltm = pd.DataFrame(ltm_data)
    headers = [i.strip() for i in headers if i is not None]
    df_ltm.columns = headers

    df_range = df_ltm.loc[(df_ltm['Periodo'] >= f'{mes_inicio[1]}-{mes_inicio[0]}-01')
                          & (df_ltm['Periodo'] <= f'{mes_fin[1]}-{mes_fin[0]}-01')]

    # # identificar meses y años de cada fila
    # meses = pd.Series(df_ltm['Periodo'].dt.month).tolist()
    # df_ltm['MES'] = meses
    #
    # yrs = pd.Series(df_ltm['Periodo'].dt.year).tolist()
    # df_ltm['YEAR'] = yrs
    #
    # mask = (int(mes_inicio) <= int(list(df_ltm['MES'])) <= int(mes_fin))
    # df_range = df.loc[mask]

    # convertir en DF_2 informacion de Totales (Totales LTM / Sntr incurrida LTM / Reserva; Totales Año Pöliza,
    # Totales YTD)

    # noinspection PyDictCreation
    totales_dict = {'Sin LTM': report_doc[str(pest)]['I19'].value, 'LTM Reserva': report_doc[str(pest)]['K19'].value,
                    'Sin Pol': report_doc[str(pest)]['I22'].value, 'Pol Reserva': report_doc[str(pest)]['K22'].value,
                    'Sin YTD': report_doc[str(pest)]['I25'].value, 'YTD Reserva': report_doc[str(pest)]['K25'].value}

    # Generación de Tablas en imagenes
    df_salud = df_range.iloc[:, [0, 2, 4, 6]]
    df_salud.iloc[:, 0] = df_salud.iloc[:, 0].dt.strftime('%m-%y')
    df_salud.iloc[:, 1] = df_salud.iloc[:, 1].apply(lambda x : f"${x:,.2f}")
    df_salud.iloc[:, 2] = df_salud.iloc[:, 2].apply(lambda x: f"${x:,.2f}")
    df_salud.iloc[:, 3] = df_salud.iloc[:, 3].apply(lambda x: f"{x*100:,.2f}%")
    df_salud.set_index('Periodo', inplace=True)

    # df_vida = df_range.iloc[:, [0, 1, 3, 5]]
    # df_vida.iloc[:, 0] = df_vida.iloc[:, 0].dt.strftime('%m-%y')

    dfi.export(df_salud,f'df_salud_{num_pol}.png')

    # generate_barchart(df_salud)
    # Generacion de Reporte en Excel usando Openpyxl

    # wb = openpyxl.Workbook()
    # sheet = wb.active
    # sheet_title = str(num_pol)
    #
    # # Write column headers
    # for col_num, column_title in enumerate(df_salud.columns, 1):
    #     sheet.cell(row=1, column=col_num, value=column_title)
    #
    # for col_num, column_tile in enumerate(df_vida.columns, 1):
    #     sheet.cell(row=1, column=col_num+len(df_salud.columns)+1, value=column_tile)
    #
    # # Write row values
    # for row_num, row_data in enumerate(df_salud.values, 2):
    #     for col_num, cell_value in enumerate(row_data, 1):
    #         sheet.cell(row=row_num, column=col_num, value=cell_value)
    #
    # val_salud = Reference(sheet, min_col=2, max_col=3, min_row=1, max_row=len(df_salud.values))
    #
    # for row_num, row_data in enumerate(df_vida.values, 2):
    #     for col_num, cell_value in enumerate(row_data, 1):
    #         sheet.cell(row=row_num, column=col_num+len(df_salud.columns)+1, value=cell_value)
    #
    # # Add Charts
    # chart_salud = openpyxl.chart.BarChart()
    # chart_salud.add_data(val_salud, titles_from_data=True)
    # chart_salud.title = f'Salud {str(num_pol)}'
    # chart_salud.set_categories(Reference(sheet, min_col=1, max_col=1, min_row=2, max_row=len(df_salud.values)))
    # chart_salud.x_axis.title = 'Periodo'
    # chart_salud.y_axis.title = 'Monto en USD'
    # sheet.add_chart(chart_salud,'A15')
    #
    # save_path = Path('./docs/sample.xlsx')
    # wb.save(save_path)

    # if num_pol in polizas:
    #     print(f'Éxito! Ese número existe; esta en la pestaña {pest}')
    #
    # else:
    #     print('Ese número de póliza no existe. Favor intentar nuevamente')
    #     return

def generate_barchart(df):
    # Create subplot and bar
    barWidth = 0.25
    fig = plt.subplots(figsize=(20, 10))

    prima_salud = df.iloc[:, 0].tolist()
    reclamo_salud = df.iloc[:,1].tolist()

    br1 = np.arange(len(prima_salud))
    br2 = [x + barWidth for x in br1]

    #Make the plot
    plt.bar(br1, prima_salud, color = 'r', width = barWidth, edgecolor = 'grey' , label = 'Prima Salud')
    plt.bar(br2, reclamo_salud, color = 'g', width = barWidth, edgecolor = 'grey' , label = 'Reclamo Salud')

    # Adding Xticks
    plt.xlabel('Periodo' , fontweight = 'bold', fontsize = 15)
    plt.ylabel('Monto $', fontweight = 'bold', fontsize = 15)
    plt.legend()
    plt.show()


def generar_reporte():
    pass
